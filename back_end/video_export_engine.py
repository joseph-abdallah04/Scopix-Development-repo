import openpyxl
from openpyxl.cell import MergedCell
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io
import os
import logging
import tempfile
from datetime import datetime
from typing import List, Dict

logger = logging.getLogger(__name__)

class VideoExportEngine:
    """
    Handles exporting video analysis results to various formats
    """
    
    def __init__(self, session_manager=None):
        self.workbook = None
        self.worksheet = None
        self.session_manager = session_manager
        self.temp_files = []  # Track temp files for cleanup

    def create_excel_export(self, frames_data: List[Dict], baseline_frame_id: str = None, 
                           export_metadata: Dict = None) -> io.BytesIO:
        """
        Create Excel file with analysis results
        
        Args:
            frames_data: List of frame data with measurements, formulas, and baseline comparisons
            baseline_frame_id: ID of the baseline frame
            export_metadata: Additional export information
            
        Returns:
            io.BytesIO: Excel file as bytes buffer
        """
        try:
            logger.info(f"Starting Excel export for {len(frames_data)} frames")
            
            self.workbook = openpyxl.Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.title = "Analysis Results"
            
            # Apply styling and structure
            self._setup_styles()
            current_row = self._add_header_section(export_metadata)
            current_row = self._add_metadata_section(frames_data, baseline_frame_id, current_row)
            
            # Process each frame
            for i, frame_data in enumerate(frames_data):
                logger.info(f"Processing frame {i+1}/{len(frames_data)}: {frame_data.get('custom_name', 'Unnamed')}")
                current_row = self._add_frame_section(frame_data, current_row)
            
            # Final formatting
            self._apply_final_formatting()
            
            # Save to memory buffer
            excel_buffer = io.BytesIO()
            self.workbook.save(excel_buffer)
            excel_buffer.seek(0)
            
            # Clean up temporary files AFTER saving Excel
            self._cleanup_temp_files()
            
            logger.info("Excel export completed successfully")
            return excel_buffer
            
        except Exception as e:
            logger.error(f"Excel export failed: {e}")
            # Clean up temporary files on error
            self._cleanup_temp_files()
            raise
    
    def _cleanup_temp_files(self):
        """Clean up all temporary files created during export"""
        for temp_file in self.temp_files:
            try:
                if os.path.exists(temp_file):
                    os.unlink(temp_file)
            except Exception as e:
                logger.warning(f"Failed to clean up temp file {temp_file}: {e}")
        self.temp_files.clear()

    def _setup_styles(self):
        """Set up Excel styles"""
        self.styles = {
            'main_header': {
                'font': Font(bold=True, size=16),
                'fill': PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'border': self._create_border()
            },
            'section_header': {
                'font': Font(bold=True, size=14),
                'fill': PatternFill(start_color="8EA9DB", end_color="8EA9DB", fill_type="solid"),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'border': self._create_border()
            },
            'sub_header': {
                'font': Font(bold=True, size=12),
                'fill': PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'border': self._create_border()
            },
            'data_cell': {
                'alignment': Alignment(horizontal='center', vertical='center'),
                'border': self._create_border()
            }
        }
    
    def _create_border(self):
        """Create standard border style"""
        return Border(
            top=Side(style='thin'),
            left=Side(style='thin'),
            bottom=Side(style='thin'),
            right=Side(style='thin')
        )
    
    def _add_header_section(self, export_metadata: Dict = None) -> int:
        """Add main header section"""
        # Main title
        self.worksheet.merge_cells('A1:H2')
        cell = self.worksheet['A1']
        cell.value = 'Video Analysis Results'
        self._apply_style(cell, 'main_header')
        
        return 4
    
    def _add_metadata_section(self, frames_data: List[Dict], baseline_frame_id: str, 
                            current_row: int) -> int:
        """Add export metadata section"""
        # Export metadata
        self.worksheet[f'A{current_row}'].value = 'Export Date:'
        self.worksheet[f'B{current_row}'].value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        current_row += 1
        
        self.worksheet[f'A{current_row}'].value = 'Total Frames:'
        self.worksheet[f'B{current_row}'].value = len(frames_data)
        current_row += 1
        
        # Find baseline frame name
        baseline_frame = next((f for f in frames_data if f.get("frame_id") == baseline_frame_id), None)
        if baseline_frame:
            self.worksheet[f'A{current_row}'].value = 'Baseline Frame:'
            baseline_name = baseline_frame.get("custom_name", f"Frame {baseline_frame.get('frame_idx')}")
            self.worksheet[f'B{current_row}'].value = baseline_name
        
        return current_row + 3
    
    def _add_frame_section(self, frame_data: Dict, current_row: int) -> int:
        """Add individual frame section with data table"""
        # Frame header
        self.worksheet.merge_cells(f'A{current_row}:H{current_row}')
        frame_name = frame_data.get('custom_name', f"Frame {frame_data.get('frame_idx')}")
        baseline_text = ' (Baseline)' if frame_data.get('is_baseline', False) else ''
        header_cell = self.worksheet[f'A{current_row}']
        header_cell.value = f"{frame_name}{baseline_text}"
        self._apply_style(header_cell, 'section_header')
        current_row += 1
        
        # Frame metadata
        self.worksheet[f'A{current_row}'].value = 'Timestamp:'
        self.worksheet[f'B{current_row}'].value = f"{frame_data.get('timestamp', 0):.3f}s"
        self.worksheet[f'C{current_row}'].value = 'Frame Index:'
        self.worksheet[f'D{current_row}'].value = frame_data.get('frame_idx', 'N/A')
        current_row += 1
        
        # Add thumbnail if available
        current_row = self._add_frame_thumbnail(frame_data, current_row)
        
        # Add measurements and formulas table
        current_row = self._add_measurements_table(frame_data, current_row)
        
        return current_row + 2  # Space between frames
    
    def _add_frame_thumbnail(self, frame_data: Dict, current_row: int) -> int:
        """Add frame thumbnail if available"""
        try:
            frame_id = frame_data.get('frame_id')
            
            if frame_id and self.session_manager:
                logger.info(f"Attempting to add thumbnail for frame {frame_id}")
                
                try:
                    # Get thumbnail bytes directly from session manager
                    thumbnail_bytes = self.session_manager.get_frame_thumbnail(frame_id)
                    
                    if thumbnail_bytes:
                        logger.info(f"Successfully got thumbnail for frame {frame_id} ({len(thumbnail_bytes)} bytes)")
                        
                        # Save thumbnail to temporary file
                        with tempfile.NamedTemporaryFile(delete=False, suffix='.jpg') as temp_file:
                            temp_file.write(thumbnail_bytes)
                            temp_thumbnail_path = temp_file.name
                        
                        # Add to cleanup list for later
                        self.temp_files.append(temp_thumbnail_path)
                        
                        # Insert image into Excel
                        img = ExcelImage(temp_thumbnail_path)
                        img.width = 150
                        img.height = 100
                        
                        # Place image starting from current row
                        self.worksheet.add_image(img, f'F{current_row}')
                        current_row += 8  # Give space for image
                        
                except Exception as e:
                    logger.warning(f"Failed to get thumbnail for frame {frame_id}: {e}")
                    current_row += 2
            else:
                current_row += 2
                
        except Exception as e:
            logger.error(f"Failed to add thumbnail: {e}")
            current_row += 2
        
        return current_row
    
    def _add_measurements_table(self, frame_data: Dict, current_row: int) -> int:
        """Add comprehensive measurements and formulas table"""
        # Table headers
        headers = ['Measurement/Formula', 'Value', 'Unit', '% of Baseline', '% Change from Baseline']
        for col, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=current_row, column=col)
            cell.value = header
            self._apply_style(cell, 'sub_header')
        current_row += 1
        
        measurements = frame_data.get('measurements', {})
        formulas = frame_data.get('formulas', {})
        baseline_comparisons = frame_data.get('baseline_comparisons', {})
        is_baseline = frame_data.get('is_baseline', False)
        
        # Raw Measurements Section
        current_row = self._add_section_header(current_row, "Raw Measurements")
        
        # Angle measurements
        for angle_key, label in [('angle_a', 'Angle A'), ('angle_b', 'Angle B')]:
            if measurements.get(angle_key) is not None:
                current_row = self._add_measurement_row(
                    current_row, label, measurements[angle_key], '°', 
                    baseline_comparisons.get(angle_key), is_baseline
                )
        
        # Area measurements
        for area_key, label in [
            ('area_a', 'Area A'), ('area_b', 'Area B'),
            ('area_av', 'Area AV'), ('area_bv', 'Area BV')
        ]:
            if measurements.get(area_key) is not None:
                current_row = self._add_measurement_row(
                    current_row, label, measurements[area_key], 'px²',
                    baseline_comparisons.get(area_key), is_baseline
                )
        
        # Distance measurements
        for distance_key, label in [
            ('distance_a', 'Distance A'), ('distance_c', 'Distance C'),
            ('distance_g', 'Distance G'), ('distance_h', 'Distance H')
        ]:
            if measurements.get(distance_key) is not None:
                current_row = self._add_measurement_row(
                    current_row, label, measurements[distance_key], 'px',
                    baseline_comparisons.get(distance_key), is_baseline
                )
        
        # Calculated Formulas Section
        current_row = self._add_section_header(current_row, "Calculated Formulas")
        
        # P-Factor and C-Factor
        if formulas.get('p_factor') is not None:
            current_row = self._add_measurement_row(
                current_row, 'P-Factor (Area A / Distance A)', formulas['p_factor'], '',
                baseline_comparisons.get('p_factor'), is_baseline
            )
        
        if formulas.get('c_factor') is not None:
            current_row = self._add_measurement_row(
                current_row, 'C-Factor (Area BV / (Area AV + Area BV))', formulas['c_factor'], '',
                baseline_comparisons.get('c_factor'), is_baseline
            )
        
        # Distance Ratios
        for ratio_key, label in [
            ('distance_ratio_1', 'Distance Ratio 1 (Distance G / Distance A)'),
            ('distance_ratio_2', 'Distance Ratio 2 (Distance G / Distance C)'),
            ('distance_ratio_3', 'Distance Ratio 3 (Distance H / Distance A)'),
            ('distance_ratio_4', 'Distance Ratio 4 (Distance H / Distance C)')
        ]:
            if formulas.get(ratio_key) is not None:
                current_row = self._add_measurement_row(
                    current_row, label, formulas[ratio_key], '',
                    baseline_comparisons.get(ratio_key), is_baseline
                )
        
        # Supraglottic Area Ratios (NEW)
        if formulas.get('supraglottic_area_ratio_1') is not None:
            current_row = self._add_measurement_row(
                current_row, 'Supraglottic Area Ratio 1 (Area B / Distance A)', 
                formulas['supraglottic_area_ratio_1'], '',
                baseline_comparisons.get('supraglottic_area_ratio_1'), is_baseline
            )
        
        if formulas.get('supraglottic_area_ratio_2') is not None:
            current_row = self._add_measurement_row(
                current_row, 'Supraglottic Area Ratio 2 (Area B / (Distance A + Distance C))', 
                formulas['supraglottic_area_ratio_2'], '',
                baseline_comparisons.get('supraglottic_area_ratio_2'), is_baseline
            )
        
        return current_row + 1
    
    def _add_section_header(self, current_row: int, section_name: str) -> int:
        """Add a section header row"""
        self.worksheet.merge_cells(f'A{current_row}:E{current_row}')
        cell = self.worksheet[f'A{current_row}']
        cell.value = section_name
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = self._create_border()
        return current_row + 1
    
    def _add_measurement_row(self, current_row: int, name: str, value, unit: str,
                           baseline_comparison: Dict = None, is_baseline: bool = False) -> int:
        """Add a measurement row to the table"""
        # Measurement name
        cell = self.worksheet.cell(row=current_row, column=1)
        cell.value = name
        self._apply_style(cell, 'data_cell')
        
        # Value - FORCE CONVERSION TO NUMBER
        cell = self.worksheet.cell(row=current_row, column=2)
        if value is not None:
            try:
                # Convert to float if it's a string, then round
                numeric_value = float(value) if isinstance(value, str) else value
                cell.value = numeric_value  # Store as actual number
                cell.number_format = '0.000'  # Format display to 3 decimal places
            except (ValueError, TypeError):
                cell.value = "Invalid number"
        else:
            cell.value = "Not calculated"
        self._apply_style(cell, 'data_cell')
        
        # Unit
        cell = self.worksheet.cell(row=current_row, column=3)
        cell.value = unit
        self._apply_style(cell, 'data_cell')
        
        # Baseline comparisons
        if is_baseline:
            # For baseline frame - show N/A instead of 100.0% and 0.0%
            cell = self.worksheet.cell(row=current_row, column=4)
            cell.value = "N/A"
            self._apply_style(cell, 'data_cell')
            
            cell = self.worksheet.cell(row=current_row, column=5)
            cell.value = "N/A"
            self._apply_style(cell, 'data_cell')
        elif baseline_comparison:
            # For non-baseline frames with baseline comparisons - FORCE NUMBER CONVERSION
            cell = self.worksheet.cell(row=current_row, column=4)
            try:
                percent_value = float(baseline_comparison['percentOfBaseline']) if isinstance(baseline_comparison['percentOfBaseline'], str) else baseline_comparison['percentOfBaseline']
                cell.value = percent_value / 100  # Store as decimal for percentage formatting
                cell.number_format = '0.0%'  # Excel percentage format
            except (ValueError, TypeError, KeyError):
                cell.value = "-"
            self._apply_style(cell, 'data_cell')
            
            cell = self.worksheet.cell(row=current_row, column=5)
            try:
                change_value = float(baseline_comparison['percentChangeFromBaseline']) if isinstance(baseline_comparison['percentChangeFromBaseline'], str) else baseline_comparison['percentChangeFromBaseline']
                cell.value = change_value / 100  # Store as decimal for percentage formatting
                cell.number_format = '+0.0%;-0.0%'  # Excel percentage format with + and - signs
            except (ValueError, TypeError, KeyError):
                cell.value = "-"
            self._apply_style(cell, 'data_cell')
        else:
            # No baseline set
            cell = self.worksheet.cell(row=current_row, column=4)
            cell.value = "-"
            self._apply_style(cell, 'data_cell')
            
            cell = self.worksheet.cell(row=current_row, column=5)
            cell.value = "-"
            self._apply_style(cell, 'data_cell')
        
        return current_row + 1
    
    def _apply_style(self, cell, style_name: str):
        """Apply style to cell"""
        if style_name in self.styles:
            style = self.styles[style_name]
            if 'font' in style:
                cell.font = style['font']
            if 'fill' in style:
                cell.fill = style['fill']
            if 'alignment' in style:
                cell.alignment = style['alignment']
            if 'border' in style:
                cell.border = style['border']
    
    def _apply_final_formatting(self):
        """Apply final formatting to the worksheet"""
        # Set specific column widths to ensure proper text fit
        column_widths = {
            'A': 45,  # Measurement/Formula names (longest: "Supraglottic Area Ratio 2...")
            'B': 15,  # Value column
            'C': 8,   # Unit column
            'D': 18,  # % of Baseline column
            'E': 25   # % Change from Baseline column
        }
        
        # Apply the specific widths
        for col_letter, width in column_widths.items():
            self.worksheet.column_dimensions[col_letter].width = width

# Convenience function
def create_excel_export(frames_data: List[Dict], baseline_frame_id: str = None, 
                       export_metadata: Dict = None, session_manager=None) -> io.BytesIO:
    """
    Convenience function to create Excel export
    """
    engine = VideoExportEngine(session_manager=session_manager)
    return engine.create_excel_export(frames_data, baseline_frame_id, export_metadata)